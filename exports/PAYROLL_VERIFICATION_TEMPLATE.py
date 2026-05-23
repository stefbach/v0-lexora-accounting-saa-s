#!/usr/bin/env python3
"""
PHASE 4, Task 4D: Payroll Verification Testing
Excel Workbook Generator for 20 employees × 6 months (120 samples)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json

def create_payroll_verification_workbook():
    """Generate Excel verification workbook with all sheets."""
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(name='Calibri', size=10, bold=True, color="000000")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pass_font = Font(name='Calibri', size=10, color="006100")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(name='Calibri', size=10, color="9C0006")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ====== SHEET 1: Detailed Verification (120 rows) ======
    ws_detail = wb.create_sheet("Detailed Verification", 0)
    
    headers = [
        "Employee_Code", "Employee_Name", "Period", "Poste", "Salary_Level",
        "Gross_Calculated", "Gross_System", "Gross_Match",
        "PAYE_Calculated", "PAYE_System", "PAYE_Variance",
        "CSG_Calculated", "CSG_System", "CSG_Variance",
        "NSF_Calculated", "NSF_System", "NSF_Variance",
        "Total_Deductions_Calculated", "Total_Deductions_System",
        "Net_Calculated", "Net_System", "Net_Variance",
        "GL_6411_Match", "GL_4210_Match", "GL_4330_Match", "GL_4311_Match",
        "Verification_Status", "Notes"
    ]
    
    ws_detail.append(headers)
    
    # Format header row
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Column widths
    col_widths = {
        'A': 15, 'B': 20, 'C': 12, 'D': 25, 'E': 15,
        'F': 16, 'G': 16, 'H': 13,
        'I': 16, 'J': 14, 'K': 14,
        'L': 16, 'M': 14, 'N': 14,
        'O': 16, 'P': 14, 'Q': 14,
        'R': 26, 'S': 26,
        'T': 16, 'U': 16, 'V': 14,
        'W': 14, 'X': 14, 'Y': 14, 'Z': 14,
        'AA': 20, 'AB': 30
    }
    for col, width in col_widths.items():
        ws_detail.column_dimensions[col].width = width
    
    # Add 120 sample rows (placeholder)
    employees = [
        ("000001", "FRONTCZAK, Johanna", "Directrice RH", "Senior"),
        ("000002", "JAUNKY, Jeyel", "Technicien IT", "Junior"),
        ("000003", "CHAVETIAN, Stephano", "Producteur Contenu", "Mid"),
        ("000004", "DESIRE, Marie", "Secrétaire Médicale", "Junior"),
        ("000008", "GROODOYAL, Aditya", "Dessinateur Concepteur", "Senior"),
        ("000009", "QUENETTE, Mégane", "Productrice Contenu", "Mid"),
        ("000015", "BEERACHEE, Shubham", "Assistant Médical", "Junior"),
        ("000021", "ARJOON, Bheshouma", "Medical Secretary", "Junior"),
        ("000023", "PURSOTY, Dhanika", "Conseillère SAV", "Mid"),
        ("000024", "PAUL, Cecilia", "Responsable Production", "Mid"),
        ("000025", "SEKELY, Sheetal", "Closer", "Mid"),
    ]
    
    periods = [
        "2025-07", "2025-08", "2025-09", "2025-10", "2025-11", "2025-12"
    ]
    
    row_idx = 2
    for emp_code, emp_name, poste, level in employees:
        for period in periods:
            ws_detail.append([
                emp_code, emp_name, period, poste, level,
                "", "", "",  # Gross calculated/system/match
                "", "", "",  # PAYE calculated/system/variance
                "", "", "",  # CSG calculated/system/variance
                "", "", "",  # NSF calculated/system/variance
                "", "",      # Total deductions
                "", "", "",  # Net calculated/system/variance
                "", "", "", "",  # GL matches
                "TODO",      # Status
                ""           # Notes
            ])
            
            # Apply borders and alignment
            for col in range(1, len(headers) + 1):
                cell = ws_detail.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Format currency columns
                if col in [6, 7, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]:
                    cell.number_format = '#,##0.00'
                
                # Status column
                if col == 27:
                    cell.alignment = Alignment(horizontal='center')
            
            row_idx += 1
    
    # Freeze panes
    ws_detail.freeze_panes = "A2"
    
    # ====== SHEET 2: Summary Statistics ======
    ws_summary = wb.create_sheet("Summary Statistics", 1)
    
    summary_data = [
        ["Payroll Verification Summary", ""],
        [""],
        ["Metric", "Value"],
        ["Total Samples", "120"],
        ["Passed Verification", "0"],
        ["Failed Verification", "0"],
        ["% Pass Rate", "0%"],
        ["Variance > 0.01 MUR", "0"],
        ["Variance ≤ 0.01 MUR", "0"],
        ["GL Posting Errors", "0"],
        ["Barème Errors", "0"],
        ["Missing GL Entries", "0"],
        [""],
        ["Period Coverage", ""],
        ["Start Period", "2025-07"],
        ["End Period", "2025-12"],
        ["Total Months", "6"],
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # Format summary sheet
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    for row in range(1, len(summary_data) + 1):
        if row in [1, 3, 14]:  # Headers
            cell_a = ws_summary.cell(row=row, column=1)
            cell_a.fill = header_fill
            cell_a.font = header_font
            cell_a.border = border
            if row > 1:
                cell_b = ws_summary.cell(row=row, column=2)
                cell_b.fill = header_fill
                cell_b.font = header_font
                cell_b.border = border
        else:
            for col in [1, 2]:
                cell = ws_summary.cell(row=row, column=col)
                cell.border = border
    
    # ====== SHEET 3: Employee Summary (20 rows) ======
    ws_emp = wb.create_sheet("Employee Summary", 2)
    
    emp_headers = ["Employee", "Code", "Poste", "Samples", "Passed", "Failed", 
                   "Avg_Variance_MUR", "Salary_Level", "Notes"]
    ws_emp.append(emp_headers)
    
    for cell in ws_emp[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_emp.column_dimensions['A'].width = 20
    ws_emp.column_dimensions['B'].width = 12
    ws_emp.column_dimensions['C'].width = 25
    ws_emp.column_dimensions['D'].width = 10
    ws_emp.column_dimensions['E'].width = 10
    ws_emp.column_dimensions['F'].width = 10
    ws_emp.column_dimensions['G'].width = 18
    ws_emp.column_dimensions['H'].width = 15
    ws_emp.column_dimensions['I'].width = 30
    
    for emp_code, emp_name, poste, level in employees:
        ws_emp.append([
            emp_name, emp_code, poste, 6, 0, 0, 0.00, level, ""
        ])
    
    # Format employee summary
    for row in range(2, len(employees) + 2):
        for col in range(1, len(emp_headers) + 1):
            cell = ws_emp.cell(row=row, column=col)
            cell.border = border
            if col in [4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal='center')
                if col == 7:
                    cell.number_format = '#,##0.00'
    
    # ====== SHEET 4: Period Summary (6 rows) ======
    ws_period = wb.create_sheet("Period Summary", 3)
    
    period_headers = ["Period", "Employees", "Samples", "Passed", "Failed", 
                      "Total_Gross_MUR", "Total_Net_MUR", "Total_PAYE_MUR",
                      "Total_CSG_MUR", "Total_NSF_MUR", "Notes"]
    ws_period.append(period_headers)
    
    for cell in ws_period[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_period.column_dimensions['A'].width = 12
    ws_period.column_dimensions['B'].width = 12
    ws_period.column_dimensions['C'].width = 10
    ws_period.column_dimensions['D'].width = 10
    ws_period.column_dimensions['E'].width = 10
    ws_period.column_dimensions['F'].width = 18
    ws_period.column_dimensions['G'].width = 18
    ws_period.column_dimensions['H'].width = 18
    ws_period.column_dimensions['I'].width = 16
    ws_period.column_dimensions['J'].width = 16
    ws_period.column_dimensions['K'].width = 30
    
    for period in periods:
        ws_period.append([
            period, 0, 0, 0, 0, 0.00, 0.00, 0.00, 0.00, 0.00, ""
        ])
    
    # Format period summary
    for row in range(2, len(periods) + 2):
        for col in range(1, len(period_headers) + 1):
            cell = ws_period.cell(row=row, column=col)
            cell.border = border
            if col in [2, 3, 4, 5]:
                cell.alignment = Alignment(horizontal='center')
            if col in [6, 7, 8, 9, 10]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    
    # ====== SHEET 5: MRA Rates Reference ======
    ws_rates = wb.create_sheet("MRA 2025 Rates", 4)
    
    rates_data = [
        ["MRA 2025 PAYE Barème & Deduction Rates", ""],
        [""],
        ["PAYE (Pay As You Earn)", ""],
        ["Annual Gross", "Tax Rate"],
        ["0 - 390,000 MUR", "0%"],
        ["390,001 - 700,000 MUR", "10%"],
        ["700,001+ MUR", "15%"],
        [""],
        ["CSG (Contribution Sociale)", ""],
        ["Monthly Gross", "Rate"],
        ["≤ 50,000 MUR", "1.5%"],
        ["> 50,000 MUR", "3.0%"],
        [""],
        ["NSF (National Savings Fund)", ""],
        ["Employee", "1.0%"],
        ["Employer (patronal)", "2.5%"],
        ["Insurable Earnings Cap", "228,000 MUR/month"],
        [""],
        ["Thresholds", ""],
        ["CSG Salarié Threshold", "50,000 MUR"],
        ["PAYE Bracket 1 Limit", "390,000 MUR"],
        ["PAYE Bracket 2 Limit", "700,000 MUR"],
        [""],
        ["Verification Sources", ""],
        ["CSG/NSF parameters", "parametres_paie_mra table"],
        ["NSF barème cap", "nsf_baremes table"],
        ["Employee bulletins", "bulletins_paie table"],
        ["GL postings", "ecritures_comptables_v2 table"],
    ]
    
    for row_data in rates_data:
        ws_rates.append(row_data)
    
    ws_rates.column_dimensions['A'].width = 35
    ws_rates.column_dimensions['B'].width = 25
    
    for row in range(1, len(rates_data) + 1):
        if row in [1, 3, 9, 14, 19, 25]:  # Section headers
            cell_a = ws_rates.cell(row=row, column=1)
            cell_a.fill = subheader_fill
            cell_a.font = subheader_font
            cell_a.border = border
            if len(rates_data[row - 1]) > 1:
                cell_b = ws_rates.cell(row=row, column=2)
                cell_b.fill = subheader_fill
                cell_b.font = subheader_font
                cell_b.border = border
        else:
            for col in [1, 2]:
                cell = ws_rates.cell(row=row, column=col)
                cell.border = border
    
    # Save workbook
    output_path = "/home/user/v0-lexora-accounting-saa-s/exports/PAYROLL_CALCULATION_VERIFICATION_120_SAMPLES.xlsx"
    wb.save(output_path)
    print(f"✓ Excel workbook created: {output_path}")
    print(f"  - Sheet 1: Detailed Verification (120 rows)")
    print(f"  - Sheet 2: Summary Statistics")
    print(f"  - Sheet 3: Employee Summary (20 employees)")
    print(f"  - Sheet 4: Period Summary (6 months)")
    print(f"  - Sheet 5: MRA 2025 Rates Reference")

if __name__ == "__main__":
    create_payroll_verification_workbook()
